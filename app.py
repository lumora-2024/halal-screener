"""
╔══════════════════════════════════════════════════════════════╗
║         🌙 HALAL STOCK SCREENER — Streamlit Web App        ║
║     Full-featured client-facing dashboard with UI          ║
╚══════════════════════════════════════════════════════════════╝

Run locally:     streamlit run app.py
Deploy:          Push to GitHub → share.streamlit.io
"""

import streamlit as st
import pandas as pd
import numpy as np
import yfinance as yf
import io
import json
import time
from datetime import datetime
from halal_screener import (
    screen_portfolio,
    fetch_stock_data,
    screen_business_activity,
    screen_financial_ratios,
    calculate_purification,
    THRESHOLDS,
    HARAM_KEYWORDS
)

# ─────────────────────────────────────────────
#  PAGE CONFIG — Must be first Streamlit call
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="🌙 Halal Stock Screener",
    page_icon="🌙",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        "Get Help":     "mailto:support@yoursite.com",
        "Report a bug": "mailto:support@yoursite.com",
        "About":        "🌙 Halal Stock Screener — Built with QuantGPT | AAOIFI Standards"
    }
)

# ─────────────────────────────────────────────
#  CUSTOM CSS — Deep Islamic luxury dark theme
# ─────────────────────────────────────────────
st.markdown("""
<style>
/* ── Google Fonts ─────────────────────────── */
@import url('https://fonts.googleapis.com/css2?family=Cinzel:wght@400;600;700&family=Crimson+Pro:ital,wght@0,300;0,400;1,300&family=JetBrains+Mono:wght@400;500&display=swap');

/* ── Root Variables ──────────────────────── */
:root {
    --gold:          #C9A84C;
    --gold-light:    #E8CC7A;
    --gold-dim:      #8B6914;
    --emerald:       #1A6B4A;
    --emerald-light: #27A86E;
    --navy:          #0A0F1E;
    --navy-mid:      #111827;
    --navy-card:     #161D2E;
    --navy-border:   #1E2D45;
    --text-primary:  #F0EBE0;
    --text-dim:      #8B9BB4;
    --red:           #C0392B;
    --red-light:     #E74C3C;
    --amber:         #D4A017;
}

/* ── Global Reset ────────────────────────── */
html, body, [data-testid="stAppViewContainer"] {
    background: var(--navy) !important;
    color: var(--text-primary) !important;
    font-family: 'Crimson Pro', Georgia, serif !important;
}

/* ── Hide Streamlit Branding ─────────────── */
#MainMenu, footer, header { visibility: hidden; }

/* ── Sidebar ─────────────────────────────── */
[data-testid="stSidebar"] {
    background: var(--navy-mid) !important;
    border-right: 1px solid var(--navy-border) !important;
}
[data-testid="stSidebar"] * { color: var(--text-primary) !important; }

/* ── Metric Cards ────────────────────────── */
[data-testid="metric-container"] {
    background: var(--navy-card) !important;
    border: 1px solid var(--navy-border) !important;
    border-radius: 12px !important;
    padding: 1rem !important;
}
[data-testid="stMetricValue"] {
    color: var(--gold) !important;
    font-family: 'Cinzel', serif !important;
    font-size: 2rem !important;
}
[data-testid="stMetricLabel"] {
    color: var(--text-dim) !important;
    font-size: 0.8rem !important;
    letter-spacing: 0.1em !important;
    text-transform: uppercase !important;
}

/* ── Buttons ─────────────────────────────── */
.stButton > button {
    background: linear-gradient(135deg, var(--gold-dim), var(--gold)) !important;
    color: var(--navy) !important;
    border: none !important;
    border-radius: 8px !important;
    font-family: 'Cinzel', serif !important;
    font-weight: 600 !important;
    letter-spacing: 0.05em !important;
    padding: 0.6rem 2rem !important;
    transition: all 0.3s ease !important;
    width: 100%;
}
.stButton > button:hover {
    background: linear-gradient(135deg, var(--gold), var(--gold-light)) !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 20px rgba(201,168,76,0.3) !important;
}

/* ── Text Input ──────────────────────────── */
.stTextInput > div > div > input,
.stTextArea > div > div > textarea {
    background: var(--navy-card) !important;
    border: 1px solid var(--navy-border) !important;
    color: var(--text-primary) !important;
    border-radius: 8px !important;
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 0.9rem !important;
}
.stTextInput > div > div > input:focus,
.stTextArea > div > div > textarea:focus {
    border-color: var(--gold-dim) !important;
    box-shadow: 0 0 0 2px rgba(201,168,76,0.15) !important;
}

/* ── Tabs ────────────────────────────────── */
[data-baseweb="tab-list"] {
    background: transparent !important;
    border-bottom: 1px solid var(--navy-border) !important;
    gap: 0 !important;
}
[data-baseweb="tab"] {
    background: transparent !important;
    color: var(--text-dim) !important;
    font-family: 'Cinzel', serif !important;
    font-size: 0.8rem !important;
    letter-spacing: 0.08em !important;
    padding: 0.8rem 1.5rem !important;
    border-bottom: 2px solid transparent !important;
}
[aria-selected="true"] {
    color: var(--gold) !important;
    border-bottom: 2px solid var(--gold) !important;
    background: transparent !important;
}

/* ── DataFrames / Tables ─────────────────── */
[data-testid="stDataFrame"] {
    border: 1px solid var(--navy-border) !important;
    border-radius: 10px !important;
    overflow: hidden !important;
}

/* ── Expanders ───────────────────────────── */
[data-testid="stExpander"] {
    background: var(--navy-card) !important;
    border: 1px solid var(--navy-border) !important;
    border-radius: 10px !important;
}

/* ── Progress Bars ───────────────────────── */
.stProgress > div > div > div {
    background: linear-gradient(90deg, var(--gold-dim), var(--gold)) !important;
}

/* ── Select Boxes ────────────────────────── */
[data-baseweb="select"] > div {
    background: var(--navy-card) !important;
    border-color: var(--navy-border) !important;
    color: var(--text-primary) !important;
}

/* ── Divider ─────────────────────────────── */
hr {
    border-color: var(--navy-border) !important;
    margin: 1.5rem 0 !important;
}

/* ── Scrollbar ───────────────────────────── */
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: var(--navy); }
::-webkit-scrollbar-thumb { background: var(--navy-border); border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: var(--gold-dim); }

/* ── Custom Card Component ───────────────── */
.halal-card {
    background: var(--navy-card);
    border: 1px solid var(--navy-border);
    border-radius: 14px;
    padding: 1.5rem;
    margin-bottom: 1rem;
    transition: border-color 0.2s ease;
}
.halal-card:hover { border-color: var(--gold-dim); }

.halal-card.pass   { border-left: 3px solid var(--emerald-light); }
.halal-card.review { border-left: 3px solid var(--amber); }
.halal-card.fail   { border-left: 3px solid var(--red-light); }

.verdict-badge {
    display: inline-block;
    padding: 0.25rem 0.9rem;
    border-radius: 999px;
    font-family: 'Cinzel', serif;
    font-size: 0.75rem;
    font-weight: 600;
    letter-spacing: 0.08em;
}
.badge-pass   { background: rgba(39,168,110,0.15); color: #27A86E; border: 1px solid rgba(39,168,110,0.3); }
.badge-review { background: rgba(212,160,23,0.15); color: #D4A017; border: 1px solid rgba(212,160,23,0.3); }
.badge-fail   { background: rgba(192,57,43,0.15);  color: #E74C3C; border: 1px solid rgba(192,57,43,0.3); }

.ratio-row {
    display: flex;
    justify-content: space-between;
    align-items: center;
    padding: 0.4rem 0;
    border-bottom: 1px solid rgba(255,255,255,0.04);
    font-size: 0.9rem;
}
.ratio-label { color: var(--text-dim); }
.ratio-value-pass { color: #27A86E; font-family: 'JetBrains Mono', monospace; font-weight: 500; }
.ratio-value-fail { color: #E74C3C; font-family: 'JetBrains Mono', monospace; font-weight: 500; }
.ratio-value-na   { color: var(--text-dim); font-family: 'JetBrains Mono', monospace; }

.section-title {
    font-family: 'Cinzel', serif;
    font-size: 0.65rem;
    letter-spacing: 0.2em;
    text-transform: uppercase;
    color: var(--gold-dim);
    margin-bottom: 0.6rem;
    padding-bottom: 0.4rem;
    border-bottom: 1px solid var(--navy-border);
}

.purify-box {
    background: rgba(201,168,76,0.07);
    border: 1px solid rgba(201,168,76,0.2);
    border-radius: 8px;
    padding: 0.8rem 1rem;
    margin-top: 0.8rem;
    font-size: 0.88rem;
    color: var(--gold-light);
}

.ticker-pill {
    display: inline-block;
    background: var(--navy-border);
    color: var(--text-dim);
    border-radius: 4px;
    padding: 0.15rem 0.5rem;
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.8rem;
    margin: 0.1rem;
    cursor: pointer;
}
.ticker-pill:hover { background: var(--gold-dim); color: white; }
</style>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
#  HEADER
# ═══════════════════════════════════════════════════════════════
def render_header():
    st.markdown("""
    <div style="text-align:center; padding: 2.5rem 0 1.5rem;">
        <div style="font-size: 2.8rem; margin-bottom: 0.4rem;">🌙</div>
        <h1 style="font-family:'Cinzel',serif; font-size:2.2rem; font-weight:700;
                   color:#C9A84C; letter-spacing:0.12em; margin:0;">
            HALAL STOCK SCREENER
        </h1>
        <p style="font-family:'Crimson Pro',serif; font-size:1.1rem; color:#8B9BB4;
                  letter-spacing:0.08em; margin-top:0.5rem;">
            Shariah-Compliant Equity Analysis · AAOIFI Standards
        </p>
        <div style="width:80px; height:1px; background:linear-gradient(90deg,transparent,#C9A84C,transparent);
                    margin: 1rem auto;"></div>
    </div>
    """, unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
#  SIDEBAR
# ═══════════════════════════════════════════════════════════════
def render_sidebar():
    with st.sidebar:
        st.markdown("""
        <div style="text-align:center; padding: 1rem 0 0.5rem;">
            <span style="font-size:1.8rem;">🌙</span>
            <p style="font-family:'Cinzel',serif; color:#C9A84C; font-size:0.9rem;
                      letter-spacing:0.1em; margin:0.3rem 0 0;">HALAL SCREENER</p>
        </div>
        """, unsafe_allow_html=True)

        st.divider()

        # ── Thresholds ────────────────────────────────────────
        st.markdown('<p class="section-title">⚙️ Screening Standards</p>', unsafe_allow_html=True)

        standard = st.selectbox(
            "Shariah Standard",
            ["AAOIFI (Default)", "Dow Jones Islamic", "S&P Shariah", "Custom"],
            help="Different scholars/indices use different thresholds"
        )

        if standard == "Dow Jones Islamic":
            st.info("DJ Islamic uses 33% for all three financial ratios.")
            debt_limit  = 33
            int_limit   = 5
            recv_limit  = 33
        elif standard == "S&P Shariah":
            st.info("S&P uses 33% debt, 5% interest, 49% receivables.")
            debt_limit  = 33
            int_limit   = 5
            recv_limit  = 49
        elif standard == "Custom":
            debt_limit  = st.slider("Max Debt / MktCap (%)", 10, 50, 33)
            int_limit   = st.slider("Max Interest Income (%)", 1, 15, 5)
            recv_limit  = st.slider("Max Receivables (%)", 20, 70, 49)
        else:  # AAOIFI Default
            debt_limit  = 33
            int_limit   = 5
            recv_limit  = 49

        # Update thresholds globally
        THRESHOLDS["max_debt_to_market_cap"]    = debt_limit / 100
        THRESHOLDS["max_interest_income_ratio"] = int_limit / 100
        THRESHOLDS["max_receivables_ratio"]     = recv_limit / 100

        st.divider()

        # ── Quick Watchlists ──────────────────────────────────
        st.markdown('<p class="section-title">📋 Quick Watchlists</p>', unsafe_allow_html=True)

        watchlists = {
            "🖥️ Big Tech":       "AAPL, MSFT, GOOGL, META, AMZN, NVDA, TSLA",
            "🏥 Healthcare":     "JNJ, PFE, ABBV, MRK, UNH, BMY, AMGN",
            "🛒 Consumer":       "WMT, COST, TGT, MCD, SBUX, PG, KO",
            "🌙 Islamic ETFs":   "SPUS, HLAL, ISDU, UMMA",
            "🏦 Banks (Test)":   "JPM, BAC, GS, WFC, C",
            "⚡ Energy":         "XOM, CVX, COP, SLB, OXY",
        }

        selected_watchlist = st.selectbox("Load a preset", ["— Select —"] + list(watchlists.keys()))

        if selected_watchlist != "— Select —":
            st.session_state["input_tickers"] = watchlists[selected_watchlist]

        st.divider()

        # ── About ─────────────────────────────────────────────
        st.markdown('<p class="section-title">ℹ️ About</p>', unsafe_allow_html=True)
        st.markdown("""
        <div style="font-size:0.82rem; color:#8B9BB4; line-height:1.6;">
        This screener applies <strong style="color:#C9A84C">AAOIFI</strong>
        Shariah standards to evaluate stocks across:<br><br>
        • Business activity analysis<br>
        • Three financial ratio tests<br>
        • Purification percentage<br><br>
        <em>⚠️ For informational purposes only. Consult a qualified Islamic
        finance scholar for authoritative rulings.</em>
        </div>
        """, unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
#  STOCK RESULT CARD
# ═══════════════════════════════════════════════════════════════
def render_result_card(r: dict):
    """Renders a styled card for a single stock result."""
    if r.get("overall") == "⚠️ ERROR":
        st.markdown(f"""
        <div class="halal-card">
            <span style="color:#E74C3C;">⚠️ Error fetching {r['ticker']}</span>
            <span style="color:#8B9BB4; font-size:0.85rem;"> — {r.get('error','Unknown error')}</span>
        </div>
        """, unsafe_allow_html=True)
        return

    compliant  = r.get("compliant")
    card_class = "pass" if compliant is True else ("review" if compliant is None else "fail")
    badge_cls  = "badge-pass" if compliant is True else ("badge-review" if compliant is None else "badge-fail")

    # ── Ratio helper ──────────────────────────────────────────
    def ratio_html(label, value, limit, suffix="%"):
        if value is None:
            return f'<div class="ratio-row"><span class="ratio-label">{label}</span><span class="ratio-value-na">N/A</span></div>'
        fail  = value > limit
        cls   = "ratio-value-fail" if fail else "ratio-value-pass"
        icon  = "✗" if fail else "✓"
        return f'<div class="ratio-row"><span class="ratio-label">{label}</span><span class="{cls}">{icon} {value:.1f}{suffix} <span style="font-size:0.75rem;color:#8B9BB4;">(max {limit:.0f}%)</span></span></div>'

    debt_limit = THRESHOLDS["max_debt_to_market_cap"] * 100
    int_limit  = THRESHOLDS["max_interest_income_ratio"] * 100
    recv_limit = THRESHOLDS["max_receivables_ratio"] * 100

    purify_html = ""
    if r.get("purification_pct", 0) > 0:
        purify_html = f"""
        <div class="purify-box">
            🤲 <strong>Purification Required:</strong> Donate {r['purification_pct']:.3f}%
            of returns to charity to purify any residual impermissible income.
        </div>"""

    price_str = f"${r['price']:.2f}" if r.get("price") else "N/A"
    pe_str    = f"{r['pe_ratio']:.1f}×" if r.get("pe_ratio") else "N/A"
    div_str   = f"{r['dividend_yield']:.2f}%" if r.get("dividend_yield") else "—"

    st.markdown(f"""
    <div class="halal-card {card_class}">
        <!-- Header Row -->
        <div style="display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:1rem;">
            <div>
                <span style="font-family:'Cinzel',serif; font-size:1.35rem;
                             color:#E8CC7A; font-weight:700;">{r['ticker']}</span>
                <span style="color:#8B9BB4; font-size:0.88rem; margin-left:0.6rem;">
                    {r.get('name','')[:40]}
                </span><br>
                <span style="font-size:0.8rem; color:#8B9BB4;">
                    {r.get('sector','N/A')} · {r.get('country','N/A')} · {r.get('market_cap','N/A')}
                </span>
            </div>
            <span class="verdict-badge {badge_cls}">{r['overall']}</span>
        </div>

        <!-- Quick Stats -->
        <div style="display:flex; gap:1.5rem; margin-bottom:1rem; flex-wrap:wrap;">
            <span style="font-size:0.85rem;">
                <span style="color:#8B9BB4;">Price</span>
                <strong style="color:#F0EBE0; margin-left:0.4rem;">{price_str}</strong>
            </span>
            <span style="font-size:0.85rem;">
                <span style="color:#8B9BB4;">P/E</span>
                <strong style="color:#F0EBE0; margin-left:0.4rem;">{pe_str}</strong>
            </span>
            <span style="font-size:0.85rem;">
                <span style="color:#8B9BB4;">Div Yield</span>
                <strong style="color:#F0EBE0; margin-left:0.4rem;">{div_str}</strong>
            </span>
        </div>

        <!-- Screens -->
        <div style="display:grid; grid-template-columns:1fr 1fr; gap:1rem;">
            <!-- Business Activity -->
            <div>
                <p class="section-title">🕌 Business Activity</p>
                <div style="font-size:0.88rem;">
                    <span style="{'color:#27A86E' if r['biz_status'] == '✅ PASS' else ('color:#D4A017' if 'REVIEW' in r['biz_status'] else 'color:#E74C3C')};">
                        {r['biz_status']}
                    </span><br>
                    <span style="color:#8B9BB4; font-size:0.82rem;">{r.get('biz_reason','')[:80]}</span>
                </div>
            </div>

            <!-- Financial Ratios -->
            <div>
                <p class="section-title">📊 Financial Ratios</p>
                {ratio_html("Debt / Mkt Cap",   r.get("debt_ratio_pct"),     debt_limit)}
                {ratio_html("Interest Income",   r.get("interest_ratio_pct"), int_limit)}
                {ratio_html("Receivables",       r.get("recv_ratio_pct"),     recv_limit)}
            </div>
        </div>

        {purify_html}
    </div>
    """, unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
#  EXPORT FUNCTIONS
# ═══════════════════════════════════════════════════════════════
def results_to_excel_bytes(results: list) -> bytes:
    """Convert results to Excel bytes for download."""
    rows = []
    for r in results:
        rows.append({
            "Ticker":             r.get("ticker"),
            "Company":            r.get("name"),
            "Sector":             r.get("sector"),
            "Country":            r.get("country"),
            "Price ($)":          r.get("price"),
            "Market Cap":         r.get("market_cap"),
            "P/E Ratio":          r.get("pe_ratio"),
            "Dividend Yield (%)": r.get("dividend_yield"),
            "Debt / MktCap (%)":  r.get("debt_ratio_pct"),
            "Interest Inc (%)":   r.get("interest_ratio_pct"),
            "Receivables (%)":    r.get("recv_ratio_pct"),
            "Purification (%)":   r.get("purification_pct"),
            "Business Screen":    r.get("biz_status"),
            "Business Reason":    r.get("biz_reason"),
            "Financial Screen":   r.get("fin_status"),
            "Financial Reason":   r.get("fin_reason"),
            "Overall Verdict":    r.get("overall"),
            "Screened At":        datetime.now().strftime("%Y-%m-%d %H:%M"),
        })

    df  = pd.DataFrame(rows)
    buf = io.BytesIO()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Halal Screening", index=False)

        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        ws = writer.sheets["Halal Screening"]

        # Header styling
        header_fill = PatternFill(start_color="0A0F1E", end_color="0A0F1E", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="C9A84C", size=11)
        for cell in ws[1]:
            cell.fill       = header_fill
            cell.font       = header_font
            cell.alignment  = Alignment(horizontal="center", vertical="center")

        # Row color coding
        green  = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        yellow = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
        red    = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

        verdict_col = df.columns.get_loc("Overall Verdict") + 1
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(rows)+1), 1):
            verdict = str(ws.cell(row=row_idx+1, column=verdict_col).value or "")
            fill    = green if "✅ HALAL" in verdict else (yellow if "REVIEW" in verdict else red)
            for cell in row:
                cell.fill = fill

        # Auto-fit columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 45)

    buf.seek(0)
    return buf.read()


def results_to_csv(results: list) -> str:
    rows = []
    for r in results:
        rows.append({
            "Ticker":     r.get("ticker"),
            "Company":    r.get("name"),
            "Sector":     r.get("sector"),
            "Price":      r.get("price"),
            "Verdict":    r.get("overall"),
            "Debt%":      r.get("debt_ratio_pct"),
            "Interest%":  r.get("interest_ratio_pct"),
            "Purify%":    r.get("purification_pct"),
        })
    return pd.DataFrame(rows).to_csv(index=False)


# ═══════════════════════════════════════════════════════════════
#  MAIN APP
# ═══════════════════════════════════════════════════════════════
def main():
    render_header()
    render_sidebar()

    # ── State Init ────────────────────────────────────────────
    if "results" not in st.session_state:
        st.session_state.results = []
    if "input_tickers" not in st.session_state:
        st.session_state.input_tickers = "AAPL, MSFT, TSLA, NVDA, JNJ, WMT, JPM, GOOGL"

    # ─────────────────────────────────────────────────────────
    #  INPUT SECTION
    # ─────────────────────────────────────────────────────────
    st.markdown('<p class="section-title">🔍 Enter Tickers to Screen</p>', unsafe_allow_html=True)

    col_input, col_btn = st.columns([4, 1])
    with col_input:
        tickers_raw = st.text_area(
            label="Tickers",
            value=st.session_state.input_tickers,
            height=80,
            label_visibility="collapsed",
            placeholder="Enter tickers separated by commas: AAPL, MSFT, TSLA, ..."
        )
    with col_btn:
        st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
        run_btn = st.button("🔍 Screen", use_container_width=True)
        clear_btn = st.button("✕ Clear", use_container_width=True)

    if clear_btn:
        st.session_state.results = []
        st.rerun()

    # ─────────────────────────────────────────────────────────
    #  RUN SCREENING
    # ─────────────────────────────────────────────────────────
    if run_btn and tickers_raw.strip():
        tickers = [t.strip().upper() for t in tickers_raw.replace("\n", ",").split(",")
                   if t.strip()]
        tickers = list(dict.fromkeys(tickers))  # deduplicate, preserve order

        if len(tickers) > 30:
            st.warning("⚠️ Maximum 30 tickers per screen. First 30 will be used.")
            tickers = tickers[:30]

        # Progress UI
        progress_bar  = st.progress(0)
        status_text   = st.empty()
        results       = []

        for i, ticker in enumerate(tickers):
            status_text.markdown(
                f'<p style="color:#8B9BB4; font-size:0.88rem; text-align:center;">'
                f'Analyzing <strong style="color:#C9A84C">{ticker}</strong> '
                f'({i+1}/{len(tickers)})...</p>',
                unsafe_allow_html=True
            )
            # Import and call screen_stock directly
            from halal_screener import screen_stock
            result = screen_stock(ticker)
            results.append(result)
            progress_bar.progress((i + 1) / len(tickers))

        progress_bar.empty()
        status_text.empty()

        # Sort results
        order = {"✅ HALAL": 0, "⚠️ NEEDS REVIEW": 1, "❌ NOT HALAL": 2, "⚠️ ERROR": 3}
        results.sort(key=lambda x: order.get(x["overall"], 99))

        st.session_state.results = results
        st.rerun()

    # ─────────────────────────────────────────────────────────
    #  RESULTS DISPLAY
    # ─────────────────────────────────────────────────────────
    results = st.session_state.results

    if not results:
        # Empty state
        st.markdown("""
        <div style="text-align:center; padding:4rem 2rem; color:#8B9BB4;">
            <div style="font-size:3rem; margin-bottom:1rem;">🌙</div>
            <p style="font-family:'Cinzel',serif; font-size:1rem; color:#C9A84C;
                      letter-spacing:0.1em;">AWAITING ANALYSIS</p>
            <p style="font-size:0.9rem; line-height:1.7; max-width:400px; margin:0 auto;">
                Enter stock tickers above and click <strong>Screen</strong> to begin
                your Shariah-compliant equity analysis.
            </p>
        </div>
        """, unsafe_allow_html=True)
        return

    # ── Summary Metrics ───────────────────────────────────────
    halal  = sum(1 for r in results if r["compliant"] is True)
    review = sum(1 for r in results if r["compliant"] is None)
    haram  = sum(1 for r in results if r["compliant"] is False)
    total  = len(results)
    halal_pct = int((halal / total * 100)) if total > 0 else 0

    st.divider()
    st.markdown('<p class="section-title">📊 Screening Summary</p>', unsafe_allow_html=True)

    m1, m2, m3, m4, m5 = st.columns(5)
    with m1: st.metric("Total Screened",    total)
    with m2: st.metric("✅ Halal",          halal,  delta=f"{halal_pct}% compliant")
    with m3: st.metric("⚠️ Needs Review",   review)
    with m4: st.metric("❌ Not Halal",       haram)
    with m5: st.metric("🕐 Screened At",    datetime.now().strftime("%H:%M"))

    st.divider()

    # ── Tabs ──────────────────────────────────────────────────
    tab1, tab2, tab3 = st.tabs(["📋  All Results", "✅  Halal Only", "📊  Data Table"])

    # ── Tab 1: All Results ────────────────────────────────────
    with tab1:
        st.markdown(f"""
        <p style="color:#8B9BB4; font-size:0.85rem; margin-bottom:1rem;">
            Showing {total} results — sorted by compliance status.
        </p>
        """, unsafe_allow_html=True)

        # Filter options
        col_f1, col_f2 = st.columns([2, 2])
        with col_f1:
            filter_opt = st.selectbox(
                "Filter by verdict",
                ["All", "✅ Halal", "⚠️ Needs Review", "❌ Not Halal"],
                key="filter_verdict"
            )
        with col_f2:
            sort_opt = st.selectbox(
                "Sort by",
                ["Compliance Status", "Ticker (A-Z)", "Debt Ratio", "Interest Ratio"],
                key="sort_opt"
            )

        # Apply filter
        filtered = results
        if filter_opt == "✅ Halal":
            filtered = [r for r in results if r["compliant"] is True]
        elif filter_opt == "⚠️ Needs Review":
            filtered = [r for r in results if r["compliant"] is None]
        elif filter_opt == "❌ Not Halal":
            filtered = [r for r in results if r["compliant"] is False]

        # Apply sort
        if sort_opt == "Ticker (A-Z)":
            filtered = sorted(filtered, key=lambda x: x["ticker"])
        elif sort_opt == "Debt Ratio":
            filtered = sorted(filtered, key=lambda x: x.get("debt_ratio_pct") or 999)
        elif sort_opt == "Interest Ratio":
            filtered = sorted(filtered, key=lambda x: x.get("interest_ratio_pct") or 999)

        for r in filtered:
            render_result_card(r)

    # ── Tab 2: Halal Only ─────────────────────────────────────
    with tab2:
        halal_results = [r for r in results if r["compliant"] is True]

        if not halal_results:
            st.markdown("""
            <div style="text-align:center; padding:3rem; color:#8B9BB4;">
                <p>No fully compliant stocks found in this screening.</p>
                <p style="font-size:0.85rem;">Try screening different tickers or adjusting
                the threshold settings in the sidebar.</p>
            </div>
            """, unsafe_allow_html=True)
        else:
            # Ticker pills
            tickers_str = "  ".join(
                f'<span class="ticker-pill">{r["ticker"]}</span>'
                for r in halal_results
            )
            st.markdown(f"""
            <div style="margin-bottom:1.5rem;">
                <p class="section-title">Compliant Tickers ({len(halal_results)})</p>
                <div>{tickers_str}</div>
            </div>
            """, unsafe_allow_html=True)

            for r in halal_results:
                render_result_card(r)

    # ── Tab 3: Data Table ─────────────────────────────────────
    with tab3:
        table_data = []
        for r in results:
            if r.get("overall") == "⚠️ ERROR":
                continue
            table_data.append({
                "Ticker":       r.get("ticker"),
                "Company":      (r.get("name") or "")[:35],
                "Sector":       (r.get("sector") or "")[:20],
                "Price":        f"${r['price']:.2f}" if r.get("price") else "N/A",
                "Market Cap":   r.get("market_cap", "N/A"),
                "Debt %":       f"{r['debt_ratio_pct']:.1f}%" if r.get("debt_ratio_pct") is not None else "N/A",
                "Interest %":   f"{r['interest_ratio_pct']:.2f}%" if r.get("interest_ratio_pct") is not None else "N/A",
                "Receivables":  f"{r['recv_ratio_pct']:.1f}%" if r.get("recv_ratio_pct") is not None else "N/A",
                "Purify %":     f"{r['purification_pct']:.3f}%" if r.get("purification_pct", 0) > 0 else "—",
                "Verdict":      r.get("overall"),
            })

        if table_data:
            df = pd.DataFrame(table_data)
            st.dataframe(
                df,
                use_container_width=True,
                hide_index=True,
                height=400
            )

    # ─────────────────────────────────────────────────────────
    #  EXPORT SECTION
    # ─────────────────────────────────────────────────────────
    st.divider()
    st.markdown('<p class="section-title">📥 Export Results</p>', unsafe_allow_html=True)

    ex1, ex2, ex3 = st.columns(3)

    ts = datetime.now().strftime("%Y%m%d_%H%M")

    with ex1:
        excel_bytes = results_to_excel_bytes(results)
        st.download_button(
            label="📊 Download Excel Report",
            data=excel_bytes,
            file_name=f"halal_screening_{ts}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    with ex2:
        csv_str = results_to_csv(results)
        st.download_button(
            label="📄 Download CSV",
            data=csv_str,
            file_name=f"halal_screening_{ts}.csv",
            mime="text/csv",
            use_container_width=True
        )

    with ex3:
        json_str = json.dumps(results, indent=2, default=str)
        st.download_button(
            label="🗂 Download JSON",
            data=json_str,
            file_name=f"halal_screening_{ts}.json",
            mime="application/json",
            use_container_width=True
        )

    # ─────────────────────────────────────────────────────────
    #  FOOTER
    # ─────────────────────────────────────────────────────────
    st.markdown("""
    <div style="text-align:center; padding:2rem 0 1rem; color:#8B9BB4; font-size:0.8rem;">
        <div style="width:60px; height:1px; background:linear-gradient(90deg,transparent,#C9A84C,transparent);
                    margin:0 auto 1rem;"></div>
        🌙 <strong style="color:#C9A84C;">Halal Stock Screener</strong> · Built with QuantGPT<br>
        <em style="font-size:0.75rem;">For informational purposes only. This does not constitute a fatwa or religious ruling.
        Consult a qualified Islamic finance scholar for authoritative opinions.</em>
    </div>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────
if __name__ == "__main__":
    main()
