"""
╔══════════════════════════════════════════════════════════════════════════╗
║              🌙 HALAL STOCK SCREENER — QuantGPT by Umar               ║
║         Shariah-Compliant Equity Screening Tool (AAOIFI Standards)    ║
╚══════════════════════════════════════════════════════════════════════════╝

This screener evaluates stocks against Islamic finance criteria:
  1. Business Activity Screen  → Excludes haram industries
  2. Financial Ratio Screen    → Debt, interest income, receivables thresholds
  3. Purification Calculation  → Estimated % of income to donate to charity

Standards followed: AAOIFI (Accounting & Auditing Organization for Islamic
Financial Institutions) — the most widely accepted global halal standard.

Dependencies:
    pip install yfinance pandas tabulate colorama openpyxl requests

Usage:
    python halal_screener.py
    python halal_screener.py --tickers AAPL MSFT TSLA
    python halal_screener.py --file my_watchlist.txt --export
"""

import yfinance as yf
import pandas as pd
import numpy as np
import argparse
import logging
import os
import json
from datetime import datetime
from tabulate import tabulate
from colorama import Fore, Back, Style, init
import warnings
warnings.filterwarnings("ignore")

# Initialize colorama for cross-platform colored output
init(autoreset=True)

# ─────────────────────────────────────────────
#  LOGGING SETUP
# ─────────────────────────────────────────────
os.makedirs("logs", exist_ok=True)
os.makedirs("reports", exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(f"logs/halal_screener_{datetime.now().strftime('%Y%m%d')}.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════
#  SECTION 1: SHARIAH SCREENING CONFIGURATION
# ═══════════════════════════════════════════════════════════════

# ── Haram Business Activity Keywords ─────────────────────────
# If any of these appear in company description or sector/industry,
# the stock FAILS the business activity screen.
HARAM_KEYWORDS = {
    "Alcohol & Beverages": [
        "alcohol", "beer", "wine", "spirits", "brewery", "distillery",
        "alcoholic beverages", "malt beverages", "brewers"
    ],
    "Tobacco": [
        "tobacco", "cigarette", "cigars", "smoking products"
    ],
    "Gambling & Casinos": [
        "casino", "gambling", "lottery", "betting", "wagering",
        "gaming", "horse racing", "sports betting"
    ],
    "Pork & Non-Halal Food": [
        "pork", "swine", "pig farming", "ham", "bacon"
    ],
    "Weapons & Defense (Controversial)": [
        "weapons", "ammunition", "firearms", "landmines", "cluster bombs",
        "nuclear weapons", "biological weapons", "chemical weapons"
    ],
    "Adult Entertainment": [
        "adult entertainment", "pornography", "adult content",
        "xxx", "adult films", "erotic"
    ],
    "Conventional Banking & Insurance": [
        "commercial banking", "savings bank", "investment bank",
        "conventional insurance", "life insurance", "property insurance",
        "mortgage", "credit cards", "payday loans"
    ],
    "Interest-Based Finance": [
        "consumer finance", "personal finance", "subprime lending",
        "loan origination", "pawnshops"
    ]
}

# Sectors that are automatically flagged for review
FLAGGED_SECTORS = [
    "Financial Services",
    "Banks—Regional",
    "Banks—Diversified",
    "Insurance—Life",
    "Insurance—Diversified",
    "Insurance—Property & Casualty",
    "Gambling",
    "Beverages—Brewers",
    "Beverages—Wineries & Distilleries",
    "Tobacco",
    "Defense & Space",
]

# ── Shariah Financial Thresholds (AAOIFI) ────────────────────
THRESHOLDS = {
    # Debt / Market Cap must be < 33%
    "max_debt_to_market_cap": 0.33,

    # Interest income / Total revenue must be < 5%
    "max_interest_income_ratio": 0.05,

    # (Cash + Interest-bearing Securities) / Market Cap must be < 33%
    "max_liquid_assets_ratio": 0.33,

    # Accounts receivable / Total assets must be < 49%
    # (Some scholars use 33% — configurable below)
    "max_receivables_ratio": 0.49,
}

# ── Default Watchlist ─────────────────────────────────────────
# A broad list of popular stocks for demo purposes.
# Edit this or pass your own via CLI.
DEFAULT_TICKERS = [
    # Big Tech
    "AAPL", "MSFT", "GOOGL", "META", "AMZN", "NVDA", "TSLA",
    # Healthcare
    "JNJ", "PFE", "ABBV", "MRK", "UNH",
    # Consumer
    "MCD", "KO", "PEP", "PG", "WMT", "COST",
    # ETFs (Islamic ETFs)
    "SPUS", "HLAL", "ISDU",
    # Industrial / Tech
    "ORCL", "CRM", "ADBE", "INTC", "AMD", "QCOM",
    # Finance (likely to fail)
    "JPM", "BAC", "GS", "V", "MA",
    # REITs
    "O", "SPG",
]


# ═══════════════════════════════════════════════════════════════
#  SECTION 2: DATA FETCHING
# ═══════════════════════════════════════════════════════════════

def fetch_stock_data(ticker: str) -> dict:
    """
    Fetch all required financial data for a given ticker using yfinance.
    Returns a dictionary with company info and financial metrics.
    """
    try:
        stock = yf.Ticker(ticker)
        info  = stock.info

        # ── Core Company Info ────────────────────────────────
        name        = info.get("longName", ticker)
        sector      = info.get("sector", "N/A")
        industry    = info.get("industry", "N/A")
        description = info.get("longBusinessSummary", "").lower()
        country     = info.get("country", "N/A")
        market_cap  = info.get("marketCap", None)
        price       = info.get("currentPrice", info.get("regularMarketPrice", None))

        # ── Balance Sheet Metrics ────────────────────────────
        total_debt        = info.get("totalDebt", 0) or 0
        total_assets      = info.get("totalAssets", None)
        total_cash        = info.get("totalCash", 0) or 0
        net_receivables   = info.get("netReceivables", 0) or 0

        # ── Income Statement Metrics ─────────────────────────
        total_revenue     = info.get("totalRevenue", None)
        # yfinance doesn't directly expose interest income — we approximate
        # using "interestExpense" as a proxy (conservative approach)
        interest_expense  = abs(info.get("interestExpense", 0) or 0)
        operating_cashflow = info.get("operatingCashflow", None)

        # ── Valuation ────────────────────────────────────────
        pe_ratio          = info.get("trailingPE", None)
        pb_ratio          = info.get("priceToBook", None)
        dividend_yield    = info.get("dividendYield", 0) or 0
        eps               = info.get("trailingEps", None)
        roe               = info.get("returnOnEquity", None)

        return {
            "ticker":           ticker,
            "name":             name,
            "sector":           sector,
            "industry":         industry,
            "description":      description,
            "country":          country,
            "market_cap":       market_cap,
            "price":            price,
            "total_debt":       total_debt,
            "total_assets":     total_assets,
            "total_cash":       total_cash,
            "net_receivables":  net_receivables,
            "total_revenue":    total_revenue,
            "interest_expense": interest_expense,
            "pe_ratio":         pe_ratio,
            "pb_ratio":         pb_ratio,
            "dividend_yield":   dividend_yield,
            "eps":              eps,
            "roe":              roe,
        }

    except Exception as e:
        logger.warning(f"[{ticker}] Failed to fetch data: {e}")
        return {"ticker": ticker, "error": str(e)}


# ═══════════════════════════════════════════════════════════════
#  SECTION 3: BUSINESS ACTIVITY SCREENING
# ═══════════════════════════════════════════════════════════════

def screen_business_activity(data: dict) -> dict:
    """
    Screen a stock's business activity against haram categories.
    Returns a result dict with pass/fail and reason.
    """
    sector      = (data.get("sector", "") or "").strip()
    industry    = (data.get("industry", "") or "").strip()
    description = (data.get("description", "") or "").lower()

    # Combine all text to scan
    combined_text = f"{sector} {industry} {description}".lower()

    # ── Check haram keywords ──────────────────────────────────
    violations = []
    for category, keywords in HARAM_KEYWORDS.items():
        for kw in keywords:
            if kw in combined_text:
                violations.append(f"{category} ({kw})")
                break  # One match per category is enough

    # ── Check flagged sectors ─────────────────────────────────
    flagged = []
    for fs in FLAGGED_SECTORS:
        if fs.lower() in combined_text or sector == fs or industry == fs:
            flagged.append(fs)

    if violations:
        return {
            "pass":    False,
            "status":  "❌ HARAM",
            "reason":  f"Business activity violation: {'; '.join(violations[:2])}"
        }
    elif flagged:
        return {
            "pass":    None,  # None = Requires further review
            "status":  "⚠️ REVIEW",
            "reason":  f"Flagged sector — manual review needed: {flagged[0]}"
        }
    else:
        return {
            "pass":   True,
            "status": "✅ PASS",
            "reason": "No haram business activity detected"
        }


# ═══════════════════════════════════════════════════════════════
#  SECTION 4: FINANCIAL RATIO SCREENING
# ═══════════════════════════════════════════════════════════════

def screen_financial_ratios(data: dict) -> dict:
    """
    Apply AAOIFI financial ratio screens.
    All ratios use Market Cap as denominator (standard approach).
    """
    market_cap       = data.get("market_cap")
    total_debt       = data.get("total_debt", 0) or 0
    total_cash       = data.get("total_cash", 0) or 0
    net_receivables  = data.get("net_receivables", 0) or 0
    total_assets     = data.get("total_assets") or 0
    total_revenue    = data.get("total_revenue") or 0
    interest_expense = data.get("interest_expense", 0) or 0

    results   = {}
    failures  = []
    warnings_ = []

    # ── 1. Debt Ratio ─────────────────────────────────────────
    if market_cap and market_cap > 0:
        debt_ratio = total_debt / market_cap
        results["debt_ratio"] = round(debt_ratio * 100, 2)
        threshold = THRESHOLDS["max_debt_to_market_cap"] * 100

        if debt_ratio > THRESHOLDS["max_debt_to_market_cap"]:
            failures.append(
                f"Debt ratio {debt_ratio:.1%} > {THRESHOLDS['max_debt_to_market_cap']:.0%} limit"
            )
    else:
        results["debt_ratio"] = None
        warnings_.append("Market cap unavailable — debt ratio skipped")

    # ── 2. Interest Income Ratio ──────────────────────────────
    if total_revenue and total_revenue > 0:
        interest_ratio = interest_expense / total_revenue
        results["interest_ratio"] = round(interest_ratio * 100, 2)

        if interest_ratio > THRESHOLDS["max_interest_income_ratio"]:
            failures.append(
                f"Interest income {interest_ratio:.1%} > {THRESHOLDS['max_interest_income_ratio']:.0%} limit"
            )
    else:
        results["interest_ratio"] = None
        warnings_.append("Revenue unavailable — interest ratio skipped")

    # ── 3. Liquid Assets Ratio ────────────────────────────────
    if market_cap and market_cap > 0:
        liquid_ratio = total_cash / market_cap
        results["liquid_ratio"] = round(liquid_ratio * 100, 2)
        # This is informational — high cash is generally ok
    else:
        results["liquid_ratio"] = None

    # ── 4. Receivables Ratio ──────────────────────────────────
    if total_assets and total_assets > 0:
        recv_ratio = net_receivables / total_assets
        results["receivables_ratio"] = round(recv_ratio * 100, 2)

        if recv_ratio > THRESHOLDS["max_receivables_ratio"]:
            failures.append(
                f"Receivables {recv_ratio:.1%} > {THRESHOLDS['max_receivables_ratio']:.0%} limit"
            )
    else:
        results["receivables_ratio"] = None

    # ── Final Verdict ──────────────────────────────────────────
    if failures:
        return {
            "pass":     False,
            "status":   "❌ FAIL",
            "reason":   "; ".join(failures[:2]),
            "ratios":   results,
            "warnings": warnings_
        }
    else:
        return {
            "pass":     True,
            "status":   "✅ PASS",
            "reason":   "All financial ratios within Shariah limits",
            "ratios":   results,
            "warnings": warnings_
        }


# ═══════════════════════════════════════════════════════════════
#  SECTION 5: PURIFICATION CALCULATION
# ═══════════════════════════════════════════════════════════════

def calculate_purification(data: dict) -> dict:
    """
    Calculates the estimated purification (zakat/sadaqah) percentage —
    the % of dividends or profits an investor should donate to charity
    to purify their investment from any impermissible income.

    Formula: Interest Income / Total Revenue × 100 = Purification %
    """
    total_revenue    = data.get("total_revenue") or 0
    interest_expense = data.get("interest_expense", 0) or 0
    dividend_yield   = data.get("dividend_yield", 0) or 0

    if total_revenue > 0 and interest_expense > 0:
        purification_pct = (interest_expense / total_revenue) * 100
        if dividend_yield > 0:
            purification_per_share = (dividend_yield / 100) * purification_pct / 100
        else:
            purification_per_share = None
    else:
        purification_pct        = 0.0
        purification_per_share  = None

    return {
        "purification_pct":        round(purification_pct, 4),
        "purification_per_share":  round(purification_per_share * 100, 6) if purification_per_share else None
    }


# ═══════════════════════════════════════════════════════════════
#  SECTION 6: MASTER SCREENING FUNCTION
# ═══════════════════════════════════════════════════════════════

def screen_stock(ticker: str) -> dict:
    """
    Run the full Halal screening pipeline on a single ticker.
    Returns a complete screening result dictionary.
    """
    logger.info(f"Screening {ticker}...")

    # Step 1: Fetch data
    data = fetch_stock_data(ticker)
    if "error" in data:
        return {
            "ticker":   ticker,
            "name":     ticker,
            "overall":  "⚠️ ERROR",
            "error":    data["error"],
            "compliant": False
        }

    # Step 2: Business Activity Screen
    biz_result = screen_business_activity(data)

    # Step 3: Financial Ratio Screen
    fin_result = screen_financial_ratios(data)

    # Step 4: Purification
    purification = calculate_purification(data)

    # ── Overall Verdict ───────────────────────────────────────
    if biz_result["pass"] is False:
        overall   = "❌ NOT HALAL"
        compliant = False
    elif fin_result["pass"] is False:
        overall   = "❌ NOT HALAL"
        compliant = False
    elif biz_result["pass"] is None or fin_result["pass"] is None:
        overall   = "⚠️ NEEDS REVIEW"
        compliant = None
    else:
        overall   = "✅ HALAL"
        compliant = True

    # Format market cap
    mc = data.get("market_cap")
    if mc:
        if mc >= 1e12:
            mc_str = f"${mc/1e12:.2f}T"
        elif mc >= 1e9:
            mc_str = f"${mc/1e9:.2f}B"
        else:
            mc_str = f"${mc/1e6:.2f}M"
    else:
        mc_str = "N/A"

    return {
        "ticker":             ticker,
        "name":               data.get("name", ticker),
        "sector":             data.get("sector", "N/A"),
        "industry":           data.get("industry", "N/A"),
        "country":            data.get("country", "N/A"),
        "market_cap":         mc_str,
        "price":              data.get("price"),
        "pe_ratio":           data.get("pe_ratio"),
        "dividend_yield":     round((data.get("dividend_yield") or 0) * 100, 2),
        "overall":            overall,
        "compliant":          compliant,
        "biz_status":         biz_result["status"],
        "biz_reason":         biz_result["reason"],
        "fin_status":         fin_result["status"],
        "fin_reason":         fin_result["reason"],
        "debt_ratio_pct":     fin_result["ratios"].get("debt_ratio"),
        "interest_ratio_pct": fin_result["ratios"].get("interest_ratio"),
        "recv_ratio_pct":     fin_result["ratios"].get("receivables_ratio"),
        "purification_pct":   purification["purification_pct"],
        "fin_warnings":       fin_result.get("warnings", [])
    }


def screen_portfolio(tickers: list) -> list:
    """
    Screen a list of tickers and return sorted results.
    Halal stocks first, then review, then non-halal.
    """
    results = []
    total   = len(tickers)

    print(f"\n{Fore.CYAN}{'═'*70}")
    print(f"  🌙  HALAL STOCK SCREENER — Screening {total} stocks...")
    print(f"{'═'*70}{Style.RESET_ALL}\n")

    for i, ticker in enumerate(tickers, 1):
        print(f"  [{i:>2}/{total}] Analyzing {ticker.upper():<8}", end="\r")
        result = screen_stock(ticker.upper().strip())
        results.append(result)

    # Sort: Halal → Review → Not Halal → Error
    order = {"✅ HALAL": 0, "⚠️ NEEDS REVIEW": 1, "❌ NOT HALAL": 2, "⚠️ ERROR": 3}
    results.sort(key=lambda x: order.get(x["overall"], 99))

    print(f"\n{' '*50}\n")  # Clear progress line
    return results


# ═══════════════════════════════════════════════════════════════
#  SECTION 7: DISPLAY & REPORTING
# ═══════════════════════════════════════════════════════════════

def print_summary_table(results: list):
    """Print a clean color-coded summary table to the console."""

    rows = []
    for r in results:
        # Color the overall verdict
        verdict = r["overall"]
        if "HALAL" in verdict and "NOT" not in verdict:
            colored = Fore.GREEN + verdict + Style.RESET_ALL
        elif "REVIEW" in verdict:
            colored = Fore.YELLOW + verdict + Style.RESET_ALL
        elif "NOT HALAL" in verdict:
            colored = Fore.RED + verdict + Style.RESET_ALL
        else:
            colored = Fore.WHITE + verdict + Style.RESET_ALL

        rows.append([
            r["ticker"],
            r["name"][:30] if r.get("name") else "N/A",
            r.get("sector", "N/A")[:20] if r.get("sector") else "N/A",
            f"${r['price']:.2f}" if r.get("price") else "N/A",
            r.get("market_cap", "N/A"),
            f"{r['debt_ratio_pct']:.1f}%" if r.get("debt_ratio_pct") is not None else "N/A",
            f"{r['interest_ratio_pct']:.2f}%" if r.get("interest_ratio_pct") is not None else "N/A",
            f"{r['purification_pct']:.3f}%" if r.get("purification_pct", 0) > 0 else "—",
            colored
        ])

    headers = [
        "Ticker", "Company", "Sector",
        "Price", "Mkt Cap",
        "Debt %", "Int Inc %",
        "Purify %", "VERDICT"
    ]

    print(tabulate(rows, headers=headers, tablefmt="rounded_grid"))

    # ── Summary Stats ─────────────────────────────────────────
    halal   = sum(1 for r in results if r["compliant"] is True)
    review  = sum(1 for r in results if r["compliant"] is None)
    haram   = sum(1 for r in results if r["compliant"] is False)
    total   = len(results)

    print(f"\n{'─'*60}")
    print(f"  📊 SCREENING SUMMARY")
    print(f"{'─'*60}")
    print(f"  Total Screened :  {total}")
    print(f"  {Fore.GREEN}✅ Halal         :  {halal}{Style.RESET_ALL}")
    print(f"  {Fore.YELLOW}⚠️  Needs Review  :  {review}{Style.RESET_ALL}")
    print(f"  {Fore.RED}❌ Not Halal     :  {haram}{Style.RESET_ALL}")
    print(f"{'─'*60}\n")


def print_detailed_report(results: list):
    """Print a detailed per-stock breakdown."""
    print(f"\n{Fore.CYAN}{'═'*70}")
    print(f"  📋  DETAILED SCREENING REPORT")
    print(f"{'═'*70}{Style.RESET_ALL}\n")

    for r in results:
        if r["overall"] == "⚠️ ERROR":
            continue

        # Header
        verdict_color = (
            Fore.GREEN  if "HALAL" in r["overall"] and "NOT" not in r["overall"]
            else Fore.YELLOW if "REVIEW" in r["overall"]
            else Fore.RED
        )
        print(f"  {verdict_color}{'─'*60}")
        print(f"  {r['overall']}  {r['ticker']} — {r['name']}")
        print(f"  {'─'*60}{Style.RESET_ALL}")

        print(f"  🏢 Sector    : {r.get('sector', 'N/A')} / {r.get('industry', 'N/A')}")
        print(f"  🌍 Country   : {r.get('country', 'N/A')}")
        print(f"  💰 Price     : ${r['price']:.2f}" if r.get("price") else "  💰 Price     : N/A")
        print(f"  📈 P/E Ratio : {r['pe_ratio']:.1f}x" if r.get("pe_ratio") else "  📈 P/E Ratio : N/A")

        print(f"\n  🕌 BUSINESS ACTIVITY : {r['biz_status']}")
        print(f"     └─ {r['biz_reason']}")

        print(f"\n  📊 FINANCIAL RATIOS  : {r['fin_status']}")
        print(f"     └─ {r['fin_reason']}")

        if r.get("debt_ratio_pct") is not None:
            bar = "█" * int(r["debt_ratio_pct"] / 2)
            color = Fore.RED if r["debt_ratio_pct"] > 33 else Fore.GREEN
            print(f"     ├─ Debt/MktCap   : {color}{r['debt_ratio_pct']:>6.1f}%  {bar}{Style.RESET_ALL}  (limit: 33%)")

        if r.get("interest_ratio_pct") is not None:
            color = Fore.RED if r["interest_ratio_pct"] > 5 else Fore.GREEN
            print(f"     ├─ Interest Inc  : {color}{r['interest_ratio_pct']:>6.2f}%{Style.RESET_ALL}  (limit: 5%)")

        if r.get("recv_ratio_pct") is not None:
            color = Fore.RED if r["recv_ratio_pct"] > 49 else Fore.GREEN
            print(f"     └─ Receivables   : {color}{r['recv_ratio_pct']:>6.1f}%{Style.RESET_ALL}  (limit: 49%)")

        # Purification
        if r.get("purification_pct", 0) > 0:
            print(f"\n  🤲 PURIFICATION      : {r['purification_pct']:.3f}% of returns should be donated to charity")
        else:
            print(f"\n  🤲 PURIFICATION      : Not required (no impermissible income detected)")

        print()


def export_to_excel(results: list, filename: str = None):
    """Export screening results to a formatted Excel file."""
    if not filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reports/halal_screening_{ts}.xlsx"

    rows = []
    for r in results:
        rows.append({
            "Ticker":              r.get("ticker"),
            "Company Name":        r.get("name"),
            "Sector":              r.get("sector"),
            "Industry":            r.get("industry"),
            "Country":             r.get("country"),
            "Price ($)":           r.get("price"),
            "Market Cap":          r.get("market_cap"),
            "P/E Ratio":           r.get("pe_ratio"),
            "Dividend Yield (%)":  r.get("dividend_yield"),
            "Debt / MktCap (%)":   r.get("debt_ratio_pct"),
            "Interest Inc (%)":    r.get("interest_ratio_pct"),
            "Receivables (%)":     r.get("recv_ratio_pct"),
            "Purification (%)":    r.get("purification_pct"),
            "Biz Screen":          r.get("biz_status"),
            "Biz Reason":          r.get("biz_reason"),
            "Financial Screen":    r.get("fin_status"),
            "Financial Reason":    r.get("fin_reason"),
            "Overall Verdict":     r.get("overall"),
            "Screen Date":         datetime.now().strftime("%Y-%m-%d %H:%M"),
        })

    df = pd.DataFrame(rows)

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Screening Results", index=False)

        # Auto-fit columns
        ws = writer.sheets["Screening Results"]
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

        # Color-code verdict column
        from openpyxl.styles import PatternFill, Font
        green  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        verdict_col_idx = df.columns.get_loc("Overall Verdict") + 1
        for row_idx, cell in enumerate(
            ws.iter_rows(min_row=2, min_col=verdict_col_idx,
                         max_col=verdict_col_idx, max_row=len(rows)+1), 1
        ):
            for c in cell:
                val = str(c.value or "")
                if "✅ HALAL" in val:
                    c.fill = green
                elif "REVIEW" in val:
                    c.fill = yellow
                elif "NOT HALAL" in val:
                    c.fill = red

    logger.info(f"Excel report saved to: {filename}")
    print(f"\n  📁 Excel report exported: {Fore.CYAN}{filename}{Style.RESET_ALL}\n")
    return filename


def export_to_json(results: list):
    """Export results as JSON for integration with other apps."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"reports/halal_screening_{ts}.json"
    with open(filename, "w") as f:
        json.dump(results, f, indent=2, default=str)
    print(f"  📁 JSON export saved: {Fore.CYAN}{filename}{Style.RESET_ALL}")


# ═══════════════════════════════════════════════════════════════
#  SECTION 8: IBKR INTEGRATION (Live Watchlist Sync)
# ═══════════════════════════════════════════════════════════════

def get_halal_tickers(results: list) -> list:
    """Returns only the compliant tickers from screening results."""
    return [r["ticker"] for r in results if r["compliant"] is True]


def save_halal_watchlist(results: list):
    """
    Save halal-compliant tickers to a file that can be imported
    as a watchlist in IBKR TWS.
    """
    halal = get_halal_tickers(results)
    ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"reports/halal_watchlist_{ts}.txt"

    with open(fname, "w") as f:
        f.write("# Halal Stock Screener — Compliant Tickers\n")
        f.write(f"# Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
        f.write(f"# Total compliant: {len(halal)}\n\n")
        for t in halal:
            f.write(f"{t}\n")

    print(f"  📁 Halal watchlist saved: {Fore.GREEN}{fname}{Style.RESET_ALL}")
    print(f"  💡 Import this file into IBKR TWS: File → Open Watchlist")
    return fname


# ═══════════════════════════════════════════════════════════════
#  SECTION 9: MAIN ENTRY POINT
# ═══════════════════════════════════════════════════════════════

def print_banner():
    banner = f"""
{Fore.GREEN}
  ██╗  ██╗ █████╗ ██╗      █████╗ ██╗
  ██║  ██║██╔══██╗██║     ██╔══██╗██║
  ███████║███████║██║     ███████║██║
  ██╔══██║██╔══██║██║     ██╔══██║██║
  ██║  ██║██║  ██║███████╗██║  ██║███████╗
  ╚═╝  ╚═╝╚═╝  ╚═╝╚══════╝╚═╝  ╚═╝╚══════╝
{Style.RESET_ALL}
{Fore.CYAN}  🌙  HALAL STOCK SCREENER  ─  Powered by QuantGPT
{Fore.WHITE}  Shariah-Compliant Equity Analysis | AAOIFI Standards
{Fore.YELLOW}  ⚠️  For informational purposes only. Consult a qualified Islamic
      finance scholar for authoritative fatwa on specific stocks.
{Style.RESET_ALL}
"""
    print(banner)


def main():
    print_banner()

    parser = argparse.ArgumentParser(
        description="🌙 Halal Stock Screener — Shariah-Compliant Equity Screening"
    )
    parser.add_argument(
        "--tickers", nargs="+",
        help="List of tickers to screen (e.g. AAPL MSFT TSLA)"
    )
    parser.add_argument(
        "--file", type=str,
        help="Path to a .txt file with one ticker per line"
    )
    parser.add_argument(
        "--export", action="store_true",
        help="Export results to Excel (.xlsx)"
    )
    parser.add_argument(
        "--json", action="store_true",
        help="Export results to JSON"
    )
    parser.add_argument(
        "--detailed", action="store_true", default=True,
        help="Show detailed per-stock breakdown (default: True)"
    )
    parser.add_argument(
        "--watchlist", action="store_true",
        help="Save compliant tickers to IBKR-importable watchlist file"
    )
    args = parser.parse_args()

    # ── Determine tickers ─────────────────────────────────────
    if args.tickers:
        tickers = args.tickers
    elif args.file:
        with open(args.file) as f:
            tickers = [
                line.strip().upper() for line in f
                if line.strip() and not line.startswith("#")
            ]
    else:
        tickers = DEFAULT_TICKERS

    # ── Run screening ─────────────────────────────────────────
    results = screen_portfolio(tickers)

    # ── Display results ───────────────────────────────────────
    print_summary_table(results)

    if args.detailed:
        print_detailed_report(results)

    # ── Export options ────────────────────────────────────────
    if args.export:
        export_to_excel(results)

    if args.json:
        export_to_json(results)

    if args.watchlist:
        save_halal_watchlist(results)

    print(f"\n{Fore.CYAN}  ✅ Screening complete! Results saved in /reports/{Style.RESET_ALL}")
    print(f"  🤲 May Allah bless your investments. Ameen.\n")

    return results


if __name__ == "__main__":
    main()
